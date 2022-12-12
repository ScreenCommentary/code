import datetime
import glob
import os
import sys

# import
import librosa  # librosa==0.9.1
import natsort
import numpy as np
import pandas as pd
import webrtcvad  # webrtcvad==2.0.10
from PyQt5.QtCore import *
from PyQt5.QtGui import QPalette
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.uic import loadUi
from media import CMultiMedia
from player import *
from EditAudio import EditAudio as Audio
import os
import sys
import datetime
import pandas as pd
from PyQt5.QtWidgets import QTableWidgetItem, QTableWidget, QPushButton
from gtts import gTTS
from ffmpeg import audio
from gtts import gTTS
from openpyxl.reader.excel import load_workbook
from pydub import AudioSegment

from EditAudio import EditAudio as Audio
from media import CMultiMedia
from player import *

QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)


#
class CWidget(QWidget):
    file_sender = pyqtSignal(object)
    speed_sender = pyqtSignal(float)
    list_add = pyqtSignal(object)

    # file_sender2 = pyqtSignal(object)
    # list_sender = pyqtSignal(object)
    def __init__(self):
        super().__init__()
        loadUi('main.ui', self)
        self.setWindowTitle("Screen Commentary Video Production Program")
        self.progressBar.setValue(0)
        # Multimedia Object
        self.mp = CMultiMedia(self, self.view)
        self.player = CPlayer(self)
        self.mediaPlayer = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        # video background color
        pal = QPalette()
        pal.setColor(QPalette.Background, Qt.black)
        self.view.setAutoFillBackground(True)
        self.view.setPalette(pal)
        self.df_list = []
        self.playlist = []
        self.selectedList = [0]
        self.playOption = QMediaPlaylist.Sequential

        # volume, slider
        self.vol.setRange(0, 100)
        self.vol.setValue(50)
        # play time
        self.duration = ''
        self.pos = ''

        # signal
        self.btn_video_add.clicked.connect(self.clickAdd)
        self.btn_script_add.clicked.connect(self.clickAddExcel)
        self.btn_play.clicked.connect(self.clickPlay)
        self.btn_stop.clicked.connect(self.clickStop)
        self.btn_pause.clicked.connect(self.clickPause)
        self.btn_forward.clicked.connect(self.clickForward)
        self.btn_prev.clicked.connect(self.clickPrev)
        # 영상 제작 버튼
        self.btn_makemovie.clicked.connect(self.makeMovie)

        # self.section_list.itemDoubleClicked.connect(self)
        # self.list.itemDoubleClicked.connect(self.dbClickList)
        self.list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tts_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.list.setColumnCount(1)
        self.list.setHorizontalHeaderLabels(['Commentary List'])
        self.tts_list.setColumnCount(1)
        self.tts_list.setHorizontalHeaderLabels(['TTS'])
        self.timeline.setColumnCount(2)
        self.timeline.setHorizontalHeaderLabels(['Start Timestamp', 'Length'])
        self.insert_TTS.setColumnCount(1)
        self.insert_TTS.setHorizontalHeaderLabels(['Blank List'])
        self.timeline.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.insert_TTS.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.vol.valueChanged.connect(self.volumeChanged)
        self.bar.sliderMoved.connect(self.barChanged)
        # tts list
        self.tts_list.doubleClicked.connect(self.playTTS)
        self.tts_list.itemSelectionChanged.connect(self.tableChanged)
        # timeline dbclick
        self.timeline.doubleClicked.connect(self.moveVideo)
        # 쓰레드 설정
        self.thread = ThreadClass(parent=self)
        self.file_sender.connect(self.thread.ToTTS2)
        self.speed_sender.connect(self.thread.speedValue)
        self.thread2 = VideoThread(parent=self)
        self.list_add.connect(self.thread2.executeVad)
        # self.thread3 = MakeMovieThread(parent=self)
        # self.file_sender2.connect(self.thread3.makeMovie)
        # self.list_sender.connect(self.thread3.listGetter)

        # spinbox- speed control
        self.speed_control.setValue(1.0)
        self.speed = 0
        self.file = ''

    def clickAdd(self):
        files, ext = QFileDialog.getOpenFileNames(self, "Open Movie", '',
                                                  'Video (*.mp4 *.mpg *.mpeg *.avi *.wma *.mka)')
        self.file = files
        if files:
            self.mp.addMedia(files)
            self.thread2.timestamp_list.connect(self.timelineListAdd)
            self.thread2.length_list.connect(self.lengthListAdd)
            self.thread2.start()
            self.list_add.emit(files)
            self.writetimeTableWidget(len(self.timeline_list))
        else:
            print("file unselected")

    def timelineListAdd(self, list):
        self.timeline_list = list

    def lengthListAdd(self, list):
        self.timeline_length_list = list

    def writetimeTableWidget(self, row):  ###
        self.timeline.setRowCount(row)
        self.timeline.setColumnCount(2)
        self.insert_TTS.setRowCount(row)
        self.insert_TTS.setColumnCount(1)
        for n, value in enumerate(self.timeline_list):  # loop over items in first column
            self.timeline.setItem(n, 0, QTableWidgetItem(str(value)))
        for n, value in enumerate(self.timeline_length_list):  # loop over items in first column
            self.timeline.setItem(n, 1, QTableWidgetItem(str(value)))
            self.insert_TTS.setItem(n, 0, QTableWidgetItem(''))
        self.timeline.resizeColumnsToContents()

    def clickAddExcel(self):
        path_dir = "../TTS/*"
        fList = glob.glob(path_dir)
        fList = [file for file in fList if file.endswith('.wav')]
        fList = natsort.natsorted(fList)
        row = len(fList)
        for i in fList:
            os.remove(i)
        file_path, ext = QFileDialog.getOpenFileName(self, '파일 열기', os.getcwd(), 'excel file (*.xls *.xlsx)')
        if file_path:
            self.df_list = self.loadData(file_path)
            for i in self.df_list:
                self.cmb.addItem(i.name)
            file = file_path
            self.ToTTS(file)
            self.initTableWidget(0)

    def clickPlay(self):
        self.mp.playMedia()

    def clickStop(self):
        self.mp.stopMedia()

    def clickPause(self):
        self.mp.pauseMedia()

    def clickForward(self):
        self.mp.forwardMedia(self.pos)

    def clickPrev(self):
        self.mp.prevMedia(self.pos)

    def dbClickList(self, item):
        row = self.list.row(item)
        self.mp.playMedia(row)

    def volumeChanged(self, vol):
        self.mp.volumeMedia(vol)

    def barChanged(self, pos):
        print(pos)
        self.mp.posMoveMedia(pos)

    def updateState(self, msg):
        self.state.setText(msg)

    def updateBar(self, duration):
        self.bar.setRange(0, duration)
        self.bar.setSingleStep(int(duration / 10))
        self.bar.setPageStep(int(duration / 10))
        self.bar.setTickInterval(int(duration / 10))
        td = datetime.timedelta(milliseconds=duration)
        stime = str(td)
        idx = stime.rfind('.')
        self.duration = stime[:idx]

    def updatePos(self, pos):
        self.pos = pos
        self.bar.setValue(pos)
        td = datetime.timedelta(milliseconds=pos)
        stime = str(td)
        idx = stime.rfind('.')
        stime = f'{stime[:idx]} / {self.duration}'
        self.playtime.setText(stime)

    def loadData(self, file_name):  ###
        df_list = []
        with pd.ExcelFile(file_name) as wb:
            for i, sn in enumerate(wb.sheet_names):
                try:
                    df = pd.read_excel(wb, sheet_name=sn)
                except Exception as e:
                    print('File read error:', e)
                else:
                    df = df.fillna(0)
                    df.name = sn
                    df_list.append(df)
        return df_list

    def initTableWidget(self, id):  ###
        # 테이블 위젯 값 쓰기
        self.list.clear()
        # select dataframe
        df = self.df_list[id];
        # table write
        col = len(df.keys())
        self.list.setColumnCount(col)
        self.list.setHorizontalHeaderLabels(df.keys())

        row = len(df.index)
        self.list.setRowCount(row)
        self.writeTableWidget(id, df, row, col)

    def writeTableWidget(self, id, df, row, col):  ###
        for r in range(row):
            for c in range(col):
                item = QTableWidgetItem(str(df.iloc[r][c]))
                self.list.setItem(r, c, item)
        self.list.resizeColumnsToContents()

    def ToTTS(self, file):  # tts 변환 부분
        load_wb = load_workbook(file, data_only=True)
        # 시트 이름으로 불러오기
        load_ws = load_wb['Sheet1']
        maxrow = load_ws.max_row
        self.progressBar.setMaximum(maxrow - 1)
        self.thread.countChanged.connect(self.onCountChanged)
        self.thread.start()
        self.speed_sender.emit(self.speed_control.value())
        self.file_sender.emit(file)
        # TTS fileList
        path_dir = "../TTS/*"
        fList = glob.glob(path_dir)
        fList = [file for file in fList if file.endswith('.wav')]
        fList = natsort.natsorted(fList)
        row = len(fList)
        self.tts_list.setRowCount(row)
        self.tts_list.setColumnCount(1)
        for i in range(0, row):
            self.tts_list.setItem(i, 0, QTableWidgetItem(fList[i]))
        self.createPlaylist()

    def onCountChanged(self, value):
        self.progressBar.setValue(value)

    def playTTS(self):
        self.player.play(self.playlist, self.selectedList[0], self.playOption)

    def createPlaylist(self):
        self.playlist.clear()
        for i in range(self.tts_list.rowCount()):
            self.playlist.append(self.tts_list.item(i, 0).text())

    def tableChanged(self):
        self.selectedList.clear()
        for item in self.tts_list.selectedIndexes():
            self.selectedList.append(item.row())

        # selectedList list 화
        self.selectedList = list(set(self.selectedList))

        # tts_list 가 있고 선택된 항목이 0이라면
        if self.tts_list.rowCount() != 0 and len(self.selectedList) == 0:
            self.selectedList.append(0)

    def closeEvent(self, QCloseEvent):
        re = QMessageBox.question(self, "종료 확인", "종료하시겠습니까?",
                                  QMessageBox.Yes | QMessageBox.No)
        # remove TTS file when program is off
        path_dir = "../TTS/*"
        fList = glob.glob(path_dir)
        fList = [file for file in fList if file.endswith('.wav')]
        fList = natsort.natsorted(fList)
        row = len(fList)
        for i in fList:
            os.remove(i)
        os.remove("mp4towav.wav")

        if re == QMessageBox.Yes:
            QCloseEvent.accept()
        else:
            QCloseEvent.ignore()

    # 영상 제작 버튼 클릭 시 내용 삽입
    def makeMovie(self):
        QMessageBox.warning(self, '경고', '영상 제작을 시작합니다. \n응답없음이 떠도 종료하지 마세요.\n확인을 누르면 진행됩니다.')

        # make input parameter (tts time list)
        input_list = []
        tts_file_list = []  # file list

        for n, value in enumerate(self.timeline_list):  # loop over items in first column
            if self.insert_TTS.item(n, 0) is None:
                self.insert_TTS.setItem(n, 0, QTableWidgetItem(''))
            item = self.insert_TTS.item(n, 0).text()
            if item != '':
                input_list.append(value)
                tts_file_list.append(item)

        obj = Audio(self.file[0], input_list)
        obj.setVideo(obj.videoName,
                     obj.setAudio(obj.getOriginalAudio(obj.video),
                                  obj.getTTS(tts_file_list)),
                     obj.video)

        QMessageBox.information(self, '영상 제작 완료', '제작이 완료되었습니다.')

        # self.thread2.start()
        # self.file_sender2.emit(self.file)
        # self.list_sender.emit(self.selectedList)

    def moveVideo(self):
        row = self.timeline.currentIndex().row()
        column = self.timeline.currentIndex().column()
        if column == 0:
            data = self.timeline.item(row, column).text()
            time = int(float(data) * 1000)
            self.mp.posMoveMedia(time)
        else:
            pass


class ThreadClass(QThread, QWidget):
    countChanged = pyqtSignal(int)

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self._mutex = QMutex()
        self.speed_value = 0.0

    def speedValue(self, speed):
        self.speedValue = speed

    def run(self):
        self._mutex.lock()
        self._mutex.unlock()

    def ToTTS2(self, file):
        self._mutex.lock()
        count = 0
        self.countChanged.emit(count)
        if file is None:
            self.btn_script_add.setEnabled(False)
        else:
            load_wb = load_workbook(file, data_only=True)
            # 시트 이름으로 불러오기
            load_ws = load_wb['Sheet1']
            maxrow = load_ws.max_row

            # 셀 주소로 값 출력
            for i in range(2, maxrow + 1):
                count += 1
                a = load_ws['A' + str(i)].value
                eng_wav = gTTS(a, lang='ko')
                eng_wav.save('../TTS/kor' + str(i - 1) + '.wav')
                audio.a_speed('../TTS/kor' + str(i - 1) + '.wav', self.speedValue,
                              '../TTS/kor_FAST' + str(i - 1) + '.wav')
                if os.path.exists('../TTS/kor' + str(i - 1) + '.wav'):
                    os.remove('../TTS/kor' + str(i - 1) + '.wav')
                else:
                    print("파일 존재 안함")
                self.countChanged.emit(count)
        self._mutex.unlock()
        self.quit()
        self.wait(5000)


class VideoThread(QThread, QWidget):
    file_receive = pyqtSignal(object)
    timestamp_list = pyqtSignal(list)
    length_list = pyqtSignal(list)

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self._mutex = QMutex()

    @pyqtSlot(object)
    def executeVad(self, file, pydub=None):
        # files
        src = file[0]
        dst = "mp4towav.wav"

        # convert mp4 to wav
        sound = AudioSegment.from_file(src, format="mp4")
        sound.export(dst, format="wav")

        # load data
        file_path = "mp4towav.wav"

        # load wav file (librosa)
        y, sr = librosa.load(file_path, sr=16000)
        # convert the file to int if it is in float (Py-WebRTC requirement)
        if y.dtype.kind == 'f':
            # convert to int16
            y = np.array([int(s * 32768) for s in y])
            # bound
            y[y > 32767] = 32767
            y[y < -32768] = -32768

        # define webrtcvad VAD
        vad = webrtcvad.Vad(3)  # set aggressiveness from 0 to 3

        class Frame(object):
            """Represents a "frame" of audio data."""

            def __init__(self, bytes, timestamp, duration):
                self.bytes = bytes
                self.timestamp = timestamp
                self.duration = duration

        def frame_generator(frame_duration_ms, audio, sample_rate):
            frames = []
            n = int(sample_rate * (frame_duration_ms / 1000.0) * 2)
            offset = 0
            timestamp = 0.0
            duration = (float(n) / sample_rate) / 2.0
            while offset + n < len(audio):
                frames.append(Frame(audio[offset:offset + n], timestamp, duration))
                timestamp += duration
                offset += n

            return frames

        # 10, 20, or 30
        frame_duration_ms = 20  # ms
        frames = frame_generator(frame_duration_ms, y, sr)
        not_speech_index = []
        for i, frame in enumerate(frames):
            if not vad.is_speech(frame.bytes, sr):
                # non_speech라고 인식된 인덱스 출력 및 그 프레임의 timestamp 출력
                # # 소수점이 너무 길어져서 소수점 둘째자리에서 올림
                # print(i, end=' ')
                # print(round(frame.timestamp, 2) * 2)
                # 따로 not_speech_index에 저장해놓음
                not_speech_index.append(i)

        # for i, times in enumerate(not_speech_index):
        #     print(round((frames[times].timestamp*2),3))

        # queue 사용해서 연속된 (따로 길이를 지정하지는 않음)
        # non_speech_index에서 연속이 시작하는 인덱스를 저장함
        queue = not_speech_index
        packet = []
        tmp = []
        v = queue.pop(0)
        tmp.append(v)
        # print(v)

        while len(queue) > 0:
            vv = queue.pop(0)
            # print(vv)
            if v + 1 == vv:
                tmp.append(vv)
                v = vv
            else:
                packet.append(tmp)
                tmp = []
                tmp.append(vv)
                v = vv
        packet.append(tmp)

        # 10개이상 연속된 non_speech_index (0.1초)
        end = []
        for i, con in enumerate(packet):
            if len(packet[i]) > 9:
                end.append(packet[i])
        # for i, con in enumerate(end):
        #     print(con, len(con) * 0.02)
        start = []
        start_len = []
        for i in enumerate(end):
            start.append(i[1][0])
            start_len.append(len(i[1]))
        # print(start)
        non_speech_timestamp_list = []
        non_speech_length_list = []
        # # 연속된 인덱스의 timestamp를 가져옴
        for i, times in enumerate(start):
            # 1분짜리 영상이 30까지 있는걸로 봐서 2배를 해줘서 시간을 맞춤
            if round(start_len[i] * 0.02, 25) > 0.01:
                non_speech_timestamp_list.append(round((frames[times].timestamp * 2), 3))
                non_speech_length_list.append(round(start_len[i] * 0.02, 25))
                # print("non_speech_section start timestamp:", round((frames[times].timestamp * 2), 3))
                # print("length:", round(start_len[i] * 0.02, 10))
        self.timestamp_list.emit(non_speech_timestamp_list)
        self.length_list.emit(non_speech_length_list)
        self.quit()
        self.wait(5000)
        # return non_speech_timestamp_list, non_speech_length_list


# class MakeMovieThread(QThread,QWidget):
#     def __init__(self, parent):
#         super().__init__(parent)
#         self.parent = parent
#         self._mutex = QMutex()
#         self.selectedList=''
#
#     def makeMovie(self,file):
#         file_path=file[0]
#         '''
#         time_list : 개수?
#         playlist : tts name list
#         timeline_list : time list founded
#
#         '''
#         time_list = self.selectedList
#
#         temp = Audio(file_path, time_list)
#         # temp.videoName = "prototype.mp4"
#         # 선택된 timeline 받아오기
#         temp.setVideo(temp.videoName,
#                       temp.setAudio(temp.getOriginalAudio(temp.video),
#                                     temp.getTTS(self.selectedList),
#                                     time_list),
#                       temp.video)
#         self.quit()
#         self.wait(5000)
#     def listGetter(self,list):
#         self.selectedList = list

if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = CWidget()
    w.show()
    sys.exit(app.exec_())
